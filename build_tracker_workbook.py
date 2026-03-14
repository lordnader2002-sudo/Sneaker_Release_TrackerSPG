# file: build_tracker_workbook.py

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ReleaseRow:
    release_date: date
    retail: int
    hype: str
    confidence: str
    priority: str
    brand: str
    style: str
    tags: str
    source_primary: str
    source_secondary: str
    source_count: int
    release_url: str
    release_method: str
    image_url: str
    notes: str
    hype_score: int
    confidence_score: int
    estimated_market_value: int | None
    flip_score: int | None


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("input_json", type=Path)
    p.add_argument("--changes", type=Path, default=None)
    p.add_argument("--output", type=Path, default=Path("output/monthly_tracker.xlsx"))
    return p.parse_args()


def load_json(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return [row for row in data if isinstance(row, dict)] if isinstance(data, list) else []


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def parse_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    try:
        return int(round(float(str(value).replace("$", "").replace(",", "").strip())))
    except ValueError:
        return 0


def normalize_text(value: Any) -> str:
    return " ".join(value.split()).strip() if isinstance(value, str) else ""


def normalize_row(record: dict[str, Any]) -> ReleaseRow | None:
    release_date = parse_date(record.get("releaseDate"))
    if release_date is None:
        return None

    tags_value = record.get("tags", [])
    tags = (
        ", ".join(str(x) for x in tags_value if str(x).strip())
        if isinstance(tags_value, list)
        else normalize_text(tags_value)
    )

    source_count = parse_int(record.get("matchedSources") or record.get("sourceCount") or 1)
    if source_count <= 0:
        source_count = 1

    return ReleaseRow(
        release_date=release_date,
        retail=parse_int(record.get("retailPrice")),
        hype=normalize_text(record.get("hype")) or "LOW",
        confidence=normalize_text(record.get("confidence")) or "LOW",
        priority=normalize_text(record.get("priority")) or "Low Priority",
        brand=normalize_text(record.get("brand")) or "Unknown",
        style=normalize_text(record.get("shoeName")) or "Unknown Style",
        tags=tags,
        source_primary=normalize_text(record.get("sourcePrimary")),
        source_secondary=normalize_text(record.get("sourceSecondary")),
        source_count=source_count,
        release_url=normalize_text(record.get("releaseUrl") or record.get("sourceUrl")),
        release_method=normalize_text(record.get("releaseMethod")),
        image_url=normalize_text(record.get("imageUrl")),
        notes=normalize_text(record.get("notes")),
        hype_score=parse_int(record.get("hypeScore")),
        confidence_score=parse_int(record.get("confidenceScore")),
        estimated_market_value=(
            parse_int(record.get("estimatedMarketValue"))
            if record.get("estimatedMarketValue") not in (None, "")
            else None
        ),
        flip_score=(
            int(record["flipScore"])
            if record.get("flipScore") not in (None, "")
            else None
        ),
    )


def filter_window(rows: list[ReleaseRow], days: int) -> list[ReleaseRow]:
    start = date.today()
    end = start + timedelta(days=days)
    return [
        r
        for r in sorted(rows, key=lambda x: (x.release_date, x.brand.lower(), x.style.lower()))
        if start <= r.release_date < end
    ]


def make_title(rows: list[ReleaseRow]) -> str:
    if not rows:
        return date.today().strftime("%B %Y Shoe Releases")
    first_month = rows[0].release_date.strftime("%B %Y")
    last_month = rows[-1].release_date.strftime("%B %Y")
    return f"{first_month} Shoe Releases" if first_month == last_month else f"{first_month} - {last_month} Shoe Releases"


def autosize(ws: Any) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            size = 10 if isinstance(cell.value, (datetime, date)) else len(str(cell.value))
            widths[cell.column] = max(widths.get(cell.column, 0), size)

    for col_idx, width in widths.items():
        bonus = 6 if col_idx in (11, 12, 16, 17) else 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + bonus, 70)


def apply_theme(ws: Any, title: str, end_col: int) -> None:
    dark_fill = PatternFill("solid", fgColor="0B1C3A")
    header_fill = PatternFill("solid", fgColor="102955")
    row_fill = PatternFill("solid", fgColor="173A73")
    white_bold = Font(color="FFFFFF", bold=True)
    white_font = Font(color="FFFFFF")
    thin = Side(style="thin", color="08162E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_col = get_column_letter(end_col)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].fill = dark_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Calculated Hype Level"
    ws["A2"].fill = dark_fill
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(4, ws.max_row + 1):
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Left-align long-text columns: Style (col 11), Tags (col 12), Release URL (col 16), Notes (col 17)
        for col in (11, 12, 16, 17):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")

        # Col 5 = Hype label, Col 6 = Hype Score, Col 7 = Confidence label, Col 8 = Conf Score, Col 9 = Priority
        hype_cell = ws.cell(row=row, column=5)
        hype_score_cell = ws.cell(row=row, column=6)
        confidence_cell = ws.cell(row=row, column=7)
        conf_score_cell = ws.cell(row=row, column=8)
        priority_cell = ws.cell(row=row, column=9)

        hype_value = str(hype_cell.value or "").upper()
        confidence_value = str(confidence_cell.value or "").upper()
        priority_value = str(priority_cell.value or "").upper()

        hype_color = "FF4D4D" if hype_value == "HIGH" else ("FFC000" if hype_value == "MED" else "00E676")
        conf_color = "7CFCFF" if confidence_value == "HIGH" else ("FFD966" if confidence_value == "MED" else "FFFFFF")

        hype_cell.font = Font(color=hype_color, bold=True)
        hype_score_cell.font = Font(color=hype_color)
        confidence_cell.font = Font(color=conf_color, bold=True)
        conf_score_cell.font = Font(color=conf_color)
        priority_cell.font = Font(
            color="FF8080" if priority_value == "MUST WATCH" else ("FFE699" if priority_value == "WATCH" else "FFFFFF"),
            bold=True,
        )

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"


def write_tracker_sheet(ws: Any, title: str, rows: list[ReleaseRow]) -> None:
    headers = [
        "Date",
        "Retail",
        "Market Value",
        "Flip %",
        "Hype",
        "Hype Score",
        "Confidence",
        "Conf. Score",
        "Priority",
        "Brand",
        "Style",
        "Tags",
        "Where",
        "Source Primary",
        "Source Count",
        "Release URL",
        "Notes",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)

    for row_idx, row in enumerate(rows, start=4):
        flip_str = (
            f"+{row.flip_score}%" if row.flip_score is not None and row.flip_score >= 0
            else f"{row.flip_score}%" if row.flip_score is not None
            else ""
        )
        values = [
            row.release_date,
            row.retail if row.retail else "",
            row.estimated_market_value if row.estimated_market_value is not None else "",
            flip_str,
            row.hype,
            row.hype_score,
            row.confidence,
            row.confidence_score,
            row.priority,
            row.brand,
            row.style,
            row.tags,
            row.release_method,
            row.source_primary,
            row.source_count,
            row.release_url,
            row.notes,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        ws.cell(row=row_idx, column=1).number_format = "m/d/yyyy"
        ws.cell(row=row_idx, column=2).number_format = '"$"#,##0'
        ws.cell(row=row_idx, column=3).number_format = '"$"#,##0'

    if not rows:
        ws.cell(row=4, column=1, value="No releases found for this window.")

    apply_theme(ws, title, len(headers))
    autosize(ws)


def write_changes_sheet(ws: Any, changes: list[dict[str, Any]]) -> None:
    headers = ["Change Type", "Date", "Brand", "Style", "Field Changed", "Old Value", "New Value", "Detected At"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    for row_idx, change in enumerate(changes, start=2):
        vals = [
            normalize_text(change.get("changeType")),
            normalize_text(change.get("date")),
            normalize_text(change.get("brand")),
            normalize_text(change.get("style")),
            normalize_text(change.get("fieldChanged")),
            str(change.get("oldValue", "")),
            str(change.get("newValue", "")),
            normalize_text(change.get("detectedAt")),
        ]
        for col_idx, value in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    autosize(ws)


def write_raw_sheet(ws: Any, rows: list[ReleaseRow]) -> None:
    headers = [
        "Release Date",
        "Retail",
        "Hype",
        "Confidence",
        "Priority",
        "Brand",
        "Style",
        "Tags",
        "Where",
        "Source Primary",
        "Source Secondary",
        "Source Count",
        "Release URL",
        "Image URL",
        "Notes",
        "Hype Score",
        "Confidence Score",
        "Market Value",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    for row_idx, row in enumerate(rows, start=2):
        vals = [
            row.release_date,
            row.retail,
            row.hype,
            row.confidence,
            row.priority,
            row.brand,
            row.style,
            row.tags,
            row.release_method,
            row.source_primary,
            row.source_secondary,
            row.source_count,
            row.release_url,
            row.image_url,
            row.notes,
            row.hype_score,
            row.confidence_score,
            row.estimated_market_value if row.estimated_market_value is not None else "",
        ]
        for col_idx, value in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.cell(row=row_idx, column=1).number_format = "m/d/yyyy"

    ws.freeze_panes = "A2"
    autosize(ws)


def write_high_hype_sheet(ws: Any, rows: list[ReleaseRow]) -> None:
    write_tracker_sheet(ws, "High Hype Releases", [r for r in rows if r.hype.upper() == "HIGH"])


def write_summary_sheet(ws: Any, monthly_rows: list[ReleaseRow]) -> None:
    dark_fill   = PatternFill("solid", fgColor="0B1C3A")
    header_fill = PatternFill("solid", fgColor="102955")
    row_fill    = PatternFill("solid", fgColor="173A73")
    accent_fill = PatternFill("solid", fgColor="0E2246")
    thin = Side(style="thin", color="08162E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(r: int, c: int, value: Any, fill: Any = None, font: Any = None, align: str = "left", wrap: bool = False) -> Any:
        cell = ws.cell(row=r, column=c, value=value)
        cell.fill = fill or row_fill
        cell.font = font or Font(color="FFFFFF")
        cell.border = border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        return cell

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="RELEASE SUMMARY")
    t.fill = dark_fill
    t.font = Font(color="FFFFFF", bold=True, size=16)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    sub = ws.cell(row=1, column=1)  # placeholder; actual subtitle below
    sub = ws.cell(row=2, column=1,
                  value=f"Next 35 Days  ·  {len(monthly_rows)} releases")
    sub.fill = header_fill
    sub.font = Font(color="A8C4F0", bold=False, size=11)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Computed stats ─────────────────────────────────────────────────────────
    must_watch  = sum(1 for r in monthly_rows if r.priority == "Must Watch")
    watch       = sum(1 for r in monthly_rows if r.priority == "Watch")
    high_hype   = sum(1 for r in monthly_rows if r.hype.upper() == "HIGH")
    high_conf   = sum(1 for r in monthly_rows if r.confidence.upper() == "HIGH")
    priced      = [r.retail for r in monthly_rows if r.retail > 0]
    avg_retail  = round(sum(priced) / len(priced)) if priced else None
    flippable   = [r for r in monthly_rows if r.estimated_market_value and r.retail > 0]
    avg_flip    = round(sum((r.estimated_market_value - r.retail) / r.retail * 100 for r in flippable) / len(flippable)) if flippable else None

    by_brand: dict[str, int] = {}
    by_hype:  dict[str, int] = {}
    by_conf:  dict[str, int] = {}
    by_method: dict[str, int] = {}

    for r in monthly_rows:
        by_brand[r.brand] = by_brand.get(r.brand, 0) + 1
        by_hype[r.hype.upper()] = by_hype.get(r.hype.upper(), 0) + 1
        by_conf[r.confidence.upper()] = by_conf.get(r.confidence.upper(), 0) + 1
        if r.release_method:
            by_method[r.release_method] = by_method.get(r.release_method, 0) + 1

    # ── Stats at a glance (row 4-11) ───────────────────────────────────────────
    ws.merge_cells("A4:D4")
    h = ws.cell(row=4, column=1, value="STATS AT A GLANCE")
    h.fill = header_fill
    h.font = Font(color="FFFFFF", bold=True, size=11)
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 22

    stats = [
        ("Total Releases",   len(monthly_rows),                           "FFFFFF"),
        ("Must Watch",       must_watch,                                   "FF8080"),
        ("Watch",            watch,                                        "FFE699"),
        ("High Hype",        high_hype,                                    "FF4D4D"),
        ("High Confidence",  high_conf,                                    "7CFCFF"),
        ("Avg Retail Price", f"${avg_retail}" if avg_retail else "N/A",   "A8C4F0"),
        ("Avg Flip %",       f"+{avg_flip}%" if avg_flip is not None else "N/A", "00E676"),
    ]
    for idx, (label, value, color) in enumerate(stats, start=5):
        _cell(idx, 1, label, font=Font(color="A8C4F0", bold=True), align="left")
        _cell(idx, 2, value, font=Font(color=color, bold=True, size=12), align="center")
        # Merge cols 3-4 as empty spacer with same fill
        ws.cell(row=idx, column=3).fill = row_fill
        ws.cell(row=idx, column=3).border = border
        ws.cell(row=idx, column=4).fill = row_fill
        ws.cell(row=idx, column=4).border = border
        ws.row_dimensions[idx].height = 20

    # ── Spacer ─────────────────────────────────────────────────────────────────
    cur = 5 + len(stats) + 1  # row after stats + blank line

    # ── Hype & Confidence side-by-side (cols A-B and C-D) ──────────────────────
    ws.merge_cells(f"A{cur}:B{cur}")
    h2 = ws.cell(row=cur, column=1, value="HYPE BREAKDOWN")
    h2.fill = header_fill
    h2.font = Font(color="FFFFFF", bold=True, size=11)
    h2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 22

    ws.merge_cells(f"C{cur}:D{cur}")
    h3 = ws.cell(row=cur, column=3, value="CONFIDENCE BREAKDOWN")
    h3.fill = header_fill
    h3.font = Font(color="FFFFFF", bold=True, size=11)
    h3.alignment = Alignment(horizontal="left", vertical="center")
    cur += 1

    hype_colors = {"HIGH": "FF4D4D", "MED": "FFC000", "LOW": "00E676"}
    conf_colors = {"HIGH": "7CFCFF", "MED": "FFD966", "LOW": "AAAAAA"}

    for label in ("HIGH", "MED", "LOW"):
        _cell(cur, 1, label, font=Font(color=hype_colors[label], bold=True), align="center")
        _cell(cur, 2, by_hype.get(label, 0), font=Font(color=hype_colors[label], bold=True, size=12), align="center")
        _cell(cur, 3, label, font=Font(color=conf_colors[label], bold=True), align="center")
        _cell(cur, 4, by_conf.get(label, 0), font=Font(color=conf_colors[label], bold=True, size=12), align="center")
        ws.row_dimensions[cur].height = 20
        cur += 1

    cur += 1  # blank spacer row

    # ── Release Method Breakdown ────────────────────────────────────────────────
    if by_method:
        ws.merge_cells(f"A{cur}:D{cur}")
        hm = ws.cell(row=cur, column=1, value="WHERE / RELEASE METHOD")
        hm.fill = header_fill
        hm.font = Font(color="FFFFFF", bold=True, size=11)
        hm.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur].height = 22
        cur += 1

        method_colors = {"App": "7CFCFF", "Raffle": "FFC000", "Online": "6EA8FF", "In-Store": "00E676"}
        for method, count in sorted(by_method.items(), key=lambda x: -x[1]):
            color = method_colors.get(method, "FFFFFF")
            _cell(cur, 1, method, font=Font(color=color, bold=True), align="left")
            _cell(cur, 2, count, font=Font(color=color, bold=True, size=12), align="center")
            ws.cell(row=cur, column=3).fill = row_fill
            ws.cell(row=cur, column=3).border = border
            ws.cell(row=cur, column=4).fill = row_fill
            ws.cell(row=cur, column=4).border = border
            ws.row_dimensions[cur].height = 20
            cur += 1
        cur += 1

    # ── Top Brands ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{cur}:D{cur}")
    hb = ws.cell(row=cur, column=1, value="TOP BRANDS")
    hb.fill = header_fill
    hb.font = Font(color="FFFFFF", bold=True, size=11)
    hb.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 22
    cur += 1

    top_brands = sorted(by_brand.items(), key=lambda x: (-x[1], x[0]))[:12]
    max_count = top_brands[0][1] if top_brands else 1
    for brand, count in top_brands:
        bar_len = max(1, round(count / max_count * 20))
        bar = "█" * bar_len
        _cell(cur, 1, brand, font=Font(color="FFFFFF", bold=True), align="left")
        _cell(cur, 2, count, font=Font(color="A8C4F0", bold=True), align="center")
        bar_cell = _cell(cur, 3, bar, fill=accent_fill, font=Font(color="1F5BD6"), align="left")
        bar_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=cur, column=4).fill = row_fill
        ws.cell(row=cur, column=4).border = border
        ws.row_dimensions[cur].height = 18
        cur += 1

    # ── Column widths ───────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14


def write_legend_sheet(ws: Any) -> None:
    dark_fill   = PatternFill("solid", fgColor="0B1C3A")
    header_fill = PatternFill("solid", fgColor="102955")
    row_fill    = PatternFill("solid", fgColor="173A73")
    white_bold  = Font(color="FFFFFF", bold=True, size=13)
    white_font  = Font(color="FFFFFF")
    thin = Side(style="thin", color="08162E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(r: int, text: str) -> None:
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.fill = header_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="left", vertical="center")

    def row(r: int, label: str, desc: str, label_color: str = "FFFFFF") -> None:
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=desc)
        for c in (a, b):
            c.fill = row_fill
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        a.font = Font(color=label_color, bold=True)
        b.font = white_font

    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="How Scores Work")
    title_cell.fill = dark_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=15)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Hype Score ──────────────────────────────────────────────────────────────
    hdr(3, "HYPE SCORE  (0–100+)")
    row(4,  "Brand Tier",      "HIGH brands (Air Jordan, Nike, Yeezy, New Balance, Adidas): +12 pts\n"
                                "MID brands (Vans, Converse, Puma, Reebok, ASICS, Saucony, Hoka, Salomon): +7 pts\n"
                                "Other brands: +3 pts")
    row(5,  "Collab",          "Any collab keyword (Travis Scott, Off-White, Supreme, Bad Bunny, Kith…): +20 pts")
    row(6,  "Limited / Rare",  "Limited, Exclusive, QS, PE, Raffle, Friends & Family, Sample…: +10 pts")
    row(7,  "Hot Model",       "Jordan 1–13, Dunk, Air Max, Kobe, Yeezy Boost, Samba, NB 550/990…: +12 pts")
    row(8,  "Retro / Heritage","Retro, OG, Original, Vintage, Heritage: +4 pts")
    row(9,  "Price Signal",    "Retail ≥ $250 OR retail ≤ $110: +3 pts")
    row(10, "Resale Ratio",    "Market ÷ Retail ≥ 2.5×: +35 pts  |  ≥ 2.0×: +28  |  ≥ 1.5×: +18  |  ≥ 1.2×: +8")
    row(11, "Resale Spread",   "Market − Retail ≥ $200: +15 pts  |  ≥ $150: +10  |  ≥ $75: +6  |  ≥ $30: +2")
    row(12, "Hype Level",      "HIGH ≥ 42 pts     MED ≥ 20 pts     LOW < 20 pts",
        label_color="FF4D4D")

    # ── Confidence Score ────────────────────────────────────────────────────────
    hdr(14, "CONFIDENCE SCORE  (0–100)")
    row(15, "Primary Source",   "Release confirmed by primary scraper (GOAT): +25 pts")
    row(16, "Secondary Source", "Confirmed by at least one additional source: +15 pts")
    row(17, "Retail Price",     "retailPrice field is populated: +15 pts")
    row(18, "Image",            "imageUrl field is populated: +10 pts")
    row(19, "Release URL",      "releaseUrl field is populated: +10 pts")
    row(20, "Source Count",     "Seen by 3+ sources: +28 pts  |  Seen by 2 sources: +18 pts")
    row(21, "Single-source cap","Releases seen by only 1 source are capped at 59 (cannot reach HIGH)")
    row(22, "Confidence Level", "HIGH ≥ 60 pts     MED ≥ 32 pts     LOW < 32 pts",
        label_color="7CFCFF")

    # ── Priority ────────────────────────────────────────────────────────────────
    hdr(24, "PRIORITY")
    row(25, "Must Watch",   "Hype = HIGH  AND  Confidence = HIGH",    label_color="FF8080")
    row(26, "Watch",        "Hype = HIGH + Confidence = MED, OR Hype = MED + Confidence = HIGH",
        label_color="FFE699")
    row(27, "Low Priority", "All other hype / confidence combinations")

    # ── Flip Score ──────────────────────────────────────────────────────────────
    hdr(29, "FLIP SCORE")
    row(30, "Formula",  "Flip % = round((Market Value − Retail Price) / Retail Price × 100)")
    row(31, "Example",  "$150 retail, $225 market  →  Flip % = +50%  (50% above retail)")
    row(32, "No value", "Shown as blank when retail or market value is missing")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    for r in range(3, 33):
        ws.row_dimensions[r].height = 40 if r in (4, 10, 11) else 24


def build_workbook(rows: list[ReleaseRow], changes: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    monthly_rows = filter_window(rows, 35)
    wb = Workbook()

    tracker = wb.active
    tracker.title = "Tracker"
    write_tracker_sheet(tracker, make_title(monthly_rows), monthly_rows)

    changes_ws = wb.create_sheet("Changes")
    write_changes_sheet(changes_ws, changes)

    raw = wb.create_sheet("Raw Data")
    write_raw_sheet(raw, rows)

    high_hype = wb.create_sheet("High Hype")
    write_high_hype_sheet(high_hype, monthly_rows)

    summary = wb.create_sheet("Summary")
    write_summary_sheet(summary, monthly_rows)

    legend = wb.create_sheet("How Scores Work")
    write_legend_sheet(legend)

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    raw_rows = load_json(args.input_json)
    change_rows = load_json(args.changes)

    rows = [r for r in (normalize_row(item) for item in raw_rows) if r]

    build_workbook(rows, change_rows, args.output)

    print(f"Loaded rows: {len(raw_rows)}")
    print(f"Normalized rows: {len(rows)}")
    print(f"Saved workbook: {args.output.resolve()}")


if __name__ == "__main__":
    main()
